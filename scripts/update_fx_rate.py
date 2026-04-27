#!/usr/bin/env python3
"""
Update MXN/USD FX rate in Pacifico_Bambu_Financial_Model_v2.xlsx.

Sources (in priority order):
  1. open.er-api.com  (free, no auth, daily updates)
  2. Yahoo Finance    (fallback, unofficial)

Usage:
  python3 update_fx_rate.py            # update from live source
  python3 update_fx_rate.py --dry-run  # show rate without writing
"""

import sys
import ssl
import json
import shutil
import argparse
from datetime import datetime
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import URLError

try:
    import certifi
    SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    SSL_CTX = ssl.create_default_context()
    SSL_CTX.check_hostname = False
    SSL_CTX.verify_mode = ssl.CERT_NONE

PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / "Pacifico_Bambu_Financial_Model_v2.xlsx"
LOG_FILE = PROJECT_ROOT / "scripts" / "fx_rate_log.txt"
BACKUP_DIR = PROJECT_ROOT / "scripts" / ".backups"


def fetch_from_er_api():
    url = "https://open.er-api.com/v6/latest/USD"
    req = Request(url, headers={"User-Agent": "PacificoBambu/1.0"})
    with urlopen(req, timeout=10, context=SSL_CTX) as r:
        data = json.loads(r.read())
    if data.get("result") != "success":
        raise ValueError(f"API returned: {data.get('result')}")
    rate = data["rates"]["MXN"]
    return float(rate), data.get("time_last_update_utc", "unknown"), "open.er-api.com"


def fetch_from_yahoo():
    url = "https://query1.finance.yahoo.com/v8/finance/chart/USDMXN=X"
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=10, context=SSL_CTX) as r:
        data = json.loads(r.read())
    rate = data["chart"]["result"][0]["meta"]["regularMarketPrice"]
    ts = data["chart"]["result"][0]["meta"].get("regularMarketTime", 0)
    return float(rate), datetime.utcfromtimestamp(ts).isoformat() + "Z", "Yahoo Finance"


def fetch_rate():
    errors = []
    for fetcher in (fetch_from_er_api, fetch_from_yahoo):
        try:
            return fetcher()
        except (URLError, ValueError, KeyError, json.JSONDecodeError) as e:
            errors.append(f"{fetcher.__name__}: {type(e).__name__}: {e}")
    raise RuntimeError("All FX sources failed:\n  " + "\n  ".join(errors))


def update_excel(rate: float):
    from openpyxl import load_workbook

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"FM_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(EXCEL_FILE, backup)

    wb = load_workbook(EXCEL_FILE, data_only=False)
    ws = wb["Assumptions"]
    old_rate = ws["B8"].value
    ws["B8"].value = round(rate, 4)
    wb.save(EXCEL_FILE)
    return old_rate, backup


def log(msg: str):
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a") as f:
        f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {msg}\n")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Show rate without writing")
    args = parser.parse_args()

    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found at {EXCEL_FILE}", file=sys.stderr)
        sys.exit(1)

    try:
        rate, ts, source = fetch_rate()
    except RuntimeError as e:
        log(f"FAILED: {e}")
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(2)

    print(f"Source:        {source}")
    print(f"Rate:          {rate:.4f} MXN/USD")
    print(f"Source time:   {ts}")

    if args.dry_run:
        print("\n[dry-run] Excel not modified.")
        return

    try:
        old_rate, backup = update_excel(rate)
    except Exception as e:
        log(f"WRITE FAILED: {e}")
        print(f"ERROR writing Excel: {e}", file=sys.stderr)
        sys.exit(3)

    print(f"\n✓ Updated Assumptions!B8: {old_rate} → {rate:.4f}")
    print(f"  Backup: {backup.relative_to(PROJECT_ROOT)}")
    log(f"OK source={source} rate={rate:.4f} (was {old_rate})")


if __name__ == "__main__":
    main()
